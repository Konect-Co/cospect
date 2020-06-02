
import csv

import librosa as lib
import numpy as np

import os
from openpyxl import load_workbook
import ast


#%%
base_path = os.path.sep + "home" + os.path.sep + "ravit" + os.path.sep\
     + "Konect-Code" + os.path.sep
#%%
#loading dataset contained in esc50.csv
with open(base_path + "ESC-50" + os.path.sep + "meta" + os.path.sep + "esc50.csv") as f:
    reader = csv.reader(f)
    dataset = list(reader)
#%%
#creating empty list to store the audiofile array
#this for loop stores each audio file in data set into an audio array and adds it to the list. Also, it checks if the category number matches 24, which is the category number for a cough.

x_data = []
y_data = []
#%%
for i in range(1, len(dataset)):
    print("row", i)
    row = dataset[i]
    
    if row[2] == '24':
        audio_data = row[0]
        full_path = base_path + "ESC-50" + os.path.sep + "audio" + os.path.sep + audio_data
        x,sr = lib.load(full_path)
        x_data.append(x.astype('float16'))
        y_data.append(1)
    else:
        audio_data = row[0]
        full_path = base_path + "ESC-50" + os.path.sep + "audio" + os.path.sep + audio_data
        x,sr = lib.load(full_path)
        x_data.append(x.astype('float16'))
        y_data.append(0)
#%%
sheet_path = base_path + os.path.sep + "cospect" + os.path.sep + "GenData"\
    + os.path.sep + "data.xlsx"
 

wb = load_workbook(sheet_path)
ws = wb.active

target_sample_rate = 44100

for row_i in list(range(7, 85+1)) + list(range(309,329+1)):
    name = str(row_i)
    print(row_i)
    
    yt_link = ws["H" + name].value
    
    file_path = base_path + os.path.sep + "cospect" + os.path.sep+ "Data"\
        + os.path.sep + "temp"
    command = "youtube-dl --extract-audio " + yt_link + " -o" + file_path + ".wav"
    os.system(command)
    
    opus_ext = os.path.exists(file_path + ".opus")
    m4a_ext = os.path.exists(file_path + ".m4a")
    
    if opus_ext:
        dl_path = file_path + ".opus"
    elif m4a_ext:
        dl_path = file_path + ".m4a"
	
    #TODO: Add try/catch to reading audio
    if opus_ext or m4a_ext:
        spliced_sound = np.asarray([])
        sound, sample_rate = lib.load(dl_path, mono=True)
        
        #resample audio segment to specified sample_rate
        curr_len = len(sound)
        sound = np.interp(np.arange(curr_len*(target_sample_rate/sample_rate)), np.arange(curr_len), sound)
        
        section = ws["I" + str(row_i)].value
        if(section != "FULL"):
            #print("file downloaded")
            section = ast.literal_eval(section) #string representation to actual list
            section.reverse() #to make intervals from lastest to earliest
            
            #print("length of sound", len(sound)/target_sample_rate)
            
            for interval in section: #want to consider the intervals from largest to smallest
                #print("ADDING")
                to_add = sound[int(interval[0]*target_sample_rate):int(interval[1]*target_sample_rate)] #changing seconds to milliseconds
                #print("length of sound", len(sound))
                #print("interval", int(interval[0]), int(interval[1]))
                spliced_sound = np.append(spliced_sound, to_add)
                #print(len(spliced_sound))
            #print("file spliced")
            os.remove(dl_path)
            print(spliced_sound)
            spliced_sound = np.asarray(spliced_sound).flatten().astype('float16')
        else:
            
            spliced_sound = sound.flatten().astype('float16')
        
        n = 110250 #number of measurements there should be in each sample

        #print("value of row_i", row_i)
        #print("sound_array LENGTH", len(spliced_sound))
        #print("value of n", n)
        #print(len(spliced_sound))
        chunks = [spliced_sound[t:t + n].flatten() for t in range(0, len(spliced_sound), n)]
        
        #print("chunks LENGTH", len(chunks))
        
        #padding last chunk to correct size
        #print(chunks[-1].shape[0])
        #print(len(chunks[-1]))
        padding = np.zeros(n-len(chunks[-1]), dtype=np.float16)
        chunks[-1] = np.append(chunks[-1], padding) #np.zeros((n-len(chunks[-1])))
        #print(chunks[-1])
        assert len(chunks[-1]) == n
        for chunk in chunks:
            
            x_data.append(chunk)
            y_data.append(1)
#%%
del dataset
#%%
x_data = np.asarray(x_data, dtype="float16")
y_data = np.asarray(y_data, dtype="bool")
#%%
print(len(x_data[10]))
#%%
assert x_data.shape[0] == y_data.shape[0]
perm = np.random.permutation(len(x_data))
#%%
x_data = x_data[perm]
y_data = y_data[perm]
#%%
test_prop = 0.10
cutoff = int(test_prop*len(x_data))

x_split = np.split(x_data, [cutoff])
y_split = np.split(y_data, [cutoff])

x_test = x_split[0]
x_train = x_split[1]
y_test = y_split[0]
y_train = y_split[1]

#%%
np.savez_compressed(base_path + "cospect" + os.path.sep + "GenData" + os.path.sep + "np_data",\
                    x_train=x_train, x_test=x_test, y_train=y_train, y_test=y_test)
#%%
#splitting into training and testing data
#x_train,x_test,y_train,y_test = train_test_split(x_data,y_data,test_size=0.10)
#%%
del x_data, y_data
