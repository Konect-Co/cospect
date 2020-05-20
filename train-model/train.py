import json
import numpy as np
import pandas as pd
import os
import librosa

from openpyxl import load_workbook

import ast
from pydub import AudioSegment
#%%
#TODO: This method is repeated in several locations, including "GenData/DataAugmentation"
#   Consolidate into one Python util file
#utility function to create spectrogram input from an audio file

def make_spectrogram(signal, sample_rate):
	resample_freq = 8000
	signal_resampled = librosa.core.resample(signal, sample_rate, resample_freq)

	stft = librosa.core.stft(signal_resampled)
	spectrogram = np.power(np.abs(stft), 0.5)
	
	spectrogram = np.swapaxes(spectrogram, 0, 1)
    
	return spectrogram

def file2spectrogram(file_path):
	signal, sample_rate = librosa.core.load(file_path)
	if signal.shape[0] == 2:
		signal = np.mean(signal, axis=0)

	spectrogram = make_spectrogram(signal, sample_rate)
	return spectrogram

def make_data():
    #df=pd.DataFrame(columns=['id','age','gender','symptoms','disease','audio','sampling rate','time'])
    
    dirs = "/home/ravit/Konect-Code".split("/")
    base_path = "/".join(dirs)
    
    sheet_path = base_path + "/cospect/GenData/data.xlsx"
    
    wb = load_workbook(sheet_path)
    ws = wb.active
    
    start_row = 14
    end_row = 20

    x_audio_data = []
    y_num_data = []

    for row_i in range(start_row, end_row+1):
        name = str(row_i)
        yt_link = ws["H" + name].value
        
        file_path = base_path + "/cospect/Data/temp"
        os.system("youtube-dl --extract-audio " + yt_link + " -o" + file_path + ".wav")
        
        opus_ext = os.path.exists(file_path + ".opus")
        m4a_ext = os.path.exists(file_path + ".m4a")
        
        if opus_ext:
            dl_path = file_path + ".opus"
        elif m4a_ext:
            dl_path = file_path + ".m4a"
	
        if opus_ext or m4a_ext:
            spliced_sound = sound = AudioSegment.from_file(dl_path)
            section = ws["I" + str(row_i)].value
            if(section != "FULL"):
                print("file downloaded")
                section = ast.literal_eval(section) #string representation to actual list
                section.reverse() #to make intervals from lastest to earliest
                
                for interval in section: #want to consider the intervals from largest to smallest
                    spliced_sound += sound[interval[0]*1000:interval[1]*1000] #changing seconds to milliseconds
                print("file spliced")
                os.remove(dl_path)
                #spliced_sound.export(dl_path)

            symptoms = ws["D" + str(row_i)].value.split("_")
            print("symptoms read")
            
            #resample audio segment to specified sample_rate
            sample_rate = 5000
            spliced_sound.set_frame_rate(sample_rate)
            
            chunk_len = 10 #seconds each chunk should be
            n = chunk_len*sample_rate
            
            #TODO: Not sure if this is being done correctly
            #Intended functionality: Split into 10 second segments each
            sound_array = spliced_sound.get_array_of_samples()
    
            print("value of row_i", row_i)
            print("sound_array LENGTH", len(sound_array))
            print("value of n", n)
            
            chunks = [sound_array[t:t + n] for t in range(0, len(sound_array), n)]
            
            print("chunks LENGTH", len(chunks))
            
            #padding last chunk to correct size
            chunks[-1] = np.append(chunks[-1], np.zeros((n-len(chunks[-1]))))
            x_audio_data += chunks
            
            symptom = 1 if "Dry" in symptoms else 0
            for _ in range(len(chunks)):
                y_num_data.append(symptom)
        
    x_audio_data = np.asarray(x_audio_data, dtype=np.float32)
    y_num_data = np.asarray(y_num_data, dtype=np.float32)
    
    return x_audio_data, y_num_data
#%%
x_audio_data, y_num_data = make_data()
print(x_audio_data.shape)
#%%
from sklearn.model_selection import train_test_split

training_x, testing_x, training_y, testing_y = train_test_split(x_audio_data, y_num_data, test_size=0.10)

#%%

"""
STFT for TF vs Librosa: For Reference

https://www.tensorflow.org/api_docs/python/tf/signal/stft
https://librosa.github.io/librosa/generated/librosa.core.stft.html

STFT parameters
- Window size - (TF: frame_length, Librosa: win_length)
- Length of windowed signal after padding - (TF: pad_end, Librosa: n_fft - little different functionality)
- Step length- (TF: frame_step, Librosa: hop_length)
- Windowing specification - (TF: window_fn, Librosa: window)

Librosa defaults (needed for tf)
- frame_length - 2048
- frame_step - frame_length/4 = 512
"""
#%%
from keras import Model
from keras.layers import Input, Conv1D, RNN, SimpleRNNCell, Dense, Activation, SimpleRNN, Lambda
import tensorflow as tf

def make_tf_spect(signals):
    frame_length = tf.constant(2048)
    fft_length = tf.constant(512)
    spectrogram_complex = tf.signal.stft(signals, frame_length, fft_length)
    spectrogram = tf.math.abs(spectrogram_complex)
    spectrogram = tf.transpose(spectrogram, perm=[0,2,1])
    return spectrogram

def get_model():
    conv_filters = 100
    conv_kernel_size = 100
    rnn_units = 100
    output_size = 1
    
    input_layer = Input(shape=(50000,), dtype=tf.float32)
    #input size is (None, 50000)

    #Lambda layer is necessary as tensorflow operations are involved
    spectrogram = Lambda(lambda x:make_tf_spect(x))(input_layer)
    
    #conv_output = Conv1D(conv_filters, conv_kernel_size, padding='causal')(input_layer)
    #causal padding not implemented for tensorflow js
    conv_output = Conv1D(conv_filters, conv_kernel_size)(spectrogram)
    #casual padding preserves time_steps dimension as the same
    	
    #rnn_output = RNN(SimpleRNNCell(rnn_units))(conv_output)
    rnn_output = SimpleRNN(rnn_units)(conv_output)
    #output is (None, rnn_units)
    	
    dense_output = Dense(output_size)(rnn_output)
    #output is (None, 1)
    
    output_layer = Activation('softmax')(dense_output)
    #output is (None, 1)

    model = Model(inputs=input_layer, outputs=output_layer)
    	
    return model
    

#%%
model = get_model()

from keras.optimizers import Adam    

def train (training_x, training_y, testing_x, testing_y):
	lr=1e-4
	decay=1e-6
	epochs=10

	opt = Adam(lr=lr, decay=decay)
	model.compile(loss='mse', optimizer=opt, metrics=['accuracy'])
	model.fit(training_x, training_y, epochs=epochs, validation_data=(testing_x, testing_y))

#%%
trained_model = train(training_x, training_y, testing_x, testing_y)
#%%
#Saving the model in Keras format
#TODO: Save only the model and weights, not optimizer state or any of that junk
"""

model.save_weights('./model_weights.h5')
with open("./model.json", 'w') as file:
	file.write(model.to_json())

model.save("./tf-model")

import tensorflow as tf
model_tf = tf.keras.models.load_model("./model.h5")
model_tf.save("./tf-model", save_format="tf")
"""