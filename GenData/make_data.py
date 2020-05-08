import os

from openpyxl import load_workbook
import numpy as np
import librosa

#function to create spectrogram input for our neural network from an audio file.
def make_spectrogram (audio_path, resample_to=8000, save=True):
	# load the signals and resample them
	signal, sample_rate = librosa.core.load(audio_path)
	if signal.shape[0] == 2:
		signal = np.mean(signal, axis=0)
	
	signal_resampled = librosa.core.resample(signal, sample_rate, resample_to)

	stft = librosa.core.stft(signal_resampled)
	spectrogram = np.power(np.abs(stft), 0.5)

	spectrogram = 1/(1 + np.exp(-spectrogram))
	
	max_length = 1970

	#pad the spectrograms
	spectrogram = np.pad(spectrogram, [[0,0],[1970-spectrogram.shape[1], 0]])
	
	return spectrogram

def get_input_data (paths, resample_to=8000, save=True):
	spectrograms = []
	for path in paths:
		spectrograms += [make_spectrogram(path, resample_to=resample_to, save=save)]
	spectrograms = np.reshape(spectrograms, [-1, 1025, 1970])
	spectrograms = np.swapaxes(spectrograms, 1, 2)
	return spectrograms

def get_files():
	wb = load_workbook('Data Table.xlsx')

	training_x_files = []
	training_y = []
	training_samples = 14

	ws = wb.active
	for i in range(training_samples):
		training_x_files += [ws['A' + str(i+2)].value + ".wav"]
		training_y += [1 if ws['C' + str(i+2)].value == "Whooping Cough" else 0]
	
	#Repeat the process for Testing Data
	testing_x_files = []
	testing_y = []
	testing_samples = 13

	ws = wb["Testing"]
	for i in range(testing_samples):
		testing_x_files += [ws['A' + str(i+2)].value + ".wav"]
		testing_y += [1 if ws['C' + str(i+2)].value == "Whooping Cough" else 0]
	
	training_y = np.asarray(training_y)
	testing_y = np.asarray(testing_y)

	return (training_x_files, training_y), (testing_x_files, testing_y)

def make_data():
	(training_x_files, training_y), (testing_x_files, testing_y) = get_files()

	os.chdir("./Data/processed_audio/train")
	training_x = get_input_data (training_x_files, save=False)

	os.chdir("../test")
	testing_x = get_input_data (testing_x_files, save=False)

	return (training_x, training_y), (testing_x, testing_y)

if __name__ == "__main__":
	make_data()
