"""
FOR RICHARD:

- Read relevant interval from each video and splice relevant intervals together. See following script for how-to

def splice_interval(name, interval_list, output_dir):
	sound = AudioSegment.from_wav(name + ".wav")
	new_sound = AudioSegment.empty()

	for interval in interval_list: #want to consider the intervals from largest to smallest
		new_sound += sound[interval[0]*1000:interval[1]*1000] #changing seconds to milliseconds

	new_sound.export(output_dir + name + ".wav")

- Write diagnosis in .json format and store in same Data directory as the audio recordings. (e.g. output corresponding to 7.wav should be 7.json)

"""
import os
from openpyxl import load_workbook

base_path = "/home/ubuntu/"
sheet_path = base_path + "cospect/data.xlsx"

def splice_interval(name, interval_list, output_dir):
        sound = AudioSegment.from_wav(name + ".wav")
        new_sound = AudioSegment.empty()

        for interval in interval_list: #want to consider the intervals from largest to smallest
                new_sound += sound[interval[0]*1000:interval[1]*1000] #changing seconds to milliseconds

        new_sound.export(output_dir + name + ".wav")

def gen_vid(name, link, section, path, splice_path):
	os.system("youtube-dl --extract-audio " + link + " -o " + name + ".wav")
	vidID = link.split("=")[1]

	os.chdir(path)

	opus_ext = os.path.exists(path + "/" + name + ".opus")
	m4a_ext = os.path.exists(path + "/" + name + ".m4a")

	filename = name
	if opus_ext:
		filename += ".opus"
	elif m4a_ext:
		filename += ".m4a"
	
	if opus_ext or m4a_ext:
		cmdstr = "ffmpeg -i \"" + filename + "\" -f wav -flags bitexact \"" + name + ".wav\""
		os.system(cmdstr)
                section = section[1:len(section)-1]
                interval_list = list()
                while section.find('[')>-1:
                        timeframe = section[section.find('[')+1:section.find(']')]
                        timeframe = timeframe.split()
                        for time in timeframe:
                                interval_list.append(time)
                                section = section[section.find(']'):]
                splice_interval(name, inverval_list, splice_path)
        #os.remove(path + "/" + filename)


def gen_vids (path):
        wb = load_workbook(sheet_path)
        ws = wb.active

        start_row = 7
        end_row = 311

        for row_i in range(start_row, end_row+1):
                name = str(row_i)
                yt_link = ws["H" + str(row_i)].value
                section = ws["I" + str(row_i)].value
                diagnosis = ws["D" + str(row_i)].value

                gen_vid(name, yt_link, section, base_path + "cospect/Data/YT-Audio/full-wav", base_path + "cospect/Data/YT-Audio/final-data")

wav_path = base_path + "cospect/Data/YT-Audio/full-wav"
gen_vids(wav_path)
